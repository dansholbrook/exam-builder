import os
import tempfile
import json
from typing import List, Optional, Dict, Any

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openai import OpenAI
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Load env vars
env_path = os.path.join(os.getcwd(), ".env")
load_dotenv(dotenv_path=env_path)
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise RuntimeError(f"OPENAI_API_KEY not set in .env at: {env_path}")

client = OpenAI(api_key=api_key)

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173","https://workbooklab.com","http://workbooklab.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class DataTable(BaseModel):
    headers: List[str]
    rows: List[List[Any]]

class SubQuestion(BaseModel):
    part: str  # "A", "B", "C", etc.
    question: str
    answer: str
    tolerance: Optional[float] = None

class Question(BaseModel):
    type: str  # "multiplechoice", "dropdown", "truefalse", "numerical", "data_table", "multipart"
    question: str
    answer: Optional[str] = None
    options: Optional[List[str]] = None
    tolerance: Optional[float] = None  # For numerical questions (e.g., 0.02 = 2%)
    data_table: Optional[DataTable] = None
    subquestions: Optional[List[SubQuestion]] = None

class QuestionList(BaseModel):
    questions: List[Question]
    exam_title: Optional[str] = "Professional Exam"  # ADDED: Custom exam title

class PromptRequest(BaseModel):
    prompt: str

class ExcelLayoutEngine:
    def __init__(self, workbook: Workbook):
        self.wb = workbook
        self.final_sheet = None
        self.solution_sheet = None
        self.current_row = 2
        self.answer_cells = []  # Track answer cell positions for grading
        
        # Styling
        self.title_font = Font(bold=True, size=14, color="1F4E79")
        self.question_font = Font(size=11, color="2F2F2F")
        self.bold_font = Font(bold=True, size=11)
        self.small_font = Font(size=10, color="666666")
        
        self.left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.table_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        self.table_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        self.header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.answer_fill = PatternFill(start_color="FFFEE9", end_color="FFFEE9", fill_type="solid")

    def setup_sheets(self):
        # Create Student sheet (student-facing) - CHANGED from "Final" to "Student"
        self.final_sheet = self.wb.active
        self.final_sheet.title = "Student"
        
        # Create Solution sheet (hidden, contains answers)
        self.solution_sheet = self.wb.create_sheet(title="Solution")
        
        # Set up column widths for better formatting
        for sheet in [self.final_sheet, self.solution_sheet]:
            sheet.column_dimensions['A'].width = 3
            sheet.column_dimensions['B'].width = 50
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
        
        # Fill all cells with white background in Student sheet
        for row in range(1, 100):  # Cover plenty of rows
            for col in range(1, 20):  # Cover plenty of columns
                cell = self.final_sheet.cell(row=row, column=col)
                cell.fill = self.white_fill
        
        # Add title to Student sheet - will be set by build_professional_workbook
        title_cell = self.final_sheet.cell(row=1, column=2, value="Exam")
        title_cell.font = self.title_font
        title_cell.alignment = self.center_align
        title_cell.fill = self.white_fill  # Keep title white
        self.final_sheet.row_dimensions[1].height = 30
        
        # Solution sheet headers
        headers = ["Question", "Student", "Solution", "Points"]
        for col, header in enumerate(headers, start=9):  # Start at column I
            cell = self.solution_sheet.cell(row=1, column=col, value=header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align

    def add_spacing(self, rows=1):
        """Add empty rows for better spacing"""
        self.current_row += rows

    def add_question_header(self, question_num: int, question_text: str):
        """Add a main question with professional formatting"""
        # Add some spacing before question
        self.add_spacing(1)
        
        # Question number and text
        question_cell = self.final_sheet.cell(
            row=self.current_row, 
            column=2, 
            value=f"{question_num}. {question_text}"
        )
        question_cell.font = self.bold_font
        question_cell.alignment = self.left_align
        question_cell.fill = self.white_fill  # White background
        self.final_sheet.row_dimensions[self.current_row].height = 20
        
        self.current_row += 1
        return self.current_row - 1

    def add_data_table(self, data_table: DataTable):
        """Add a professional data table"""
        if not data_table:
            return
            
        self.add_spacing(1)
        
        # Add headers
        for col, header in enumerate(data_table.headers, start=3):
            cell = self.final_sheet.cell(row=self.current_row, column=col, value=header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        self.current_row += 1
        
        # Add data rows
        for row_data in data_table.rows:
            for col, value in enumerate(row_data, start=3):
                cell = self.final_sheet.cell(row=self.current_row, column=col, value=value)
                cell.border = self.thin_border
                cell.fill = self.table_fill
                cell.alignment = self.center_align
            self.current_row += 1
            
        self.add_spacing(1)

    def add_subquestion(self, sub_q: SubQuestion, question_num: int) -> str:
        """Add a subquestion with answer input"""
        # Subquestion text
        sub_text = f"{sub_q.part}. {sub_q.question}"
        sub_cell = self.final_sheet.cell(row=self.current_row, column=2, value=sub_text)
        sub_cell.font = self.question_font
        sub_cell.alignment = self.left_align
        sub_cell.fill = self.white_fill  # White background
        
        # Answer input cell (strategically positioned)
        answer_col = 4  # Column D, like your manual exam
        answer_cell_addr = f"{get_column_letter(answer_col)}{self.current_row + 1}"
        answer_cell = self.final_sheet.cell(row=self.current_row + 1, column=answer_col, value="")
        answer_cell.border = self.thin_border
        answer_cell.alignment = self.center_align
        answer_cell.fill = self.answer_fill  # Cream color for student input
        
        # Track this for grading
        self.answer_cells.append({
            'question_id': f"{question_num}{sub_q.part}",
            'cell_address': answer_cell_addr,
            'correct_answer': sub_q.answer,
            'tolerance': sub_q.tolerance
        })
        
        self.current_row += 2
        self.add_spacing(1)
        
        return answer_cell_addr

    def add_simple_question(self, question: Question, question_num: int) -> str:
        """Add a simple question (multiple choice, true/false, etc.)"""
        self.add_spacing(1)
        
        # Question text
        q_cell = self.final_sheet.cell(row=self.current_row, column=2, value=f"{question_num}. {question.question}")
        q_cell.font = self.question_font
        q_cell.alignment = self.left_align
        q_cell.fill = self.white_fill  # White background
        
        self.current_row += 1
        
        # Answer cell
        answer_col = 3
        answer_cell_addr = f"{get_column_letter(answer_col)}{self.current_row}"
        answer_cell = self.final_sheet.cell(row=self.current_row, column=answer_col, value="")
        answer_cell.border = self.thin_border
        answer_cell.alignment = self.center_align
        answer_cell.fill = self.answer_fill  # Cream color for student input
        
        # Add dropdown validation if needed
        if question.type in ["dropdown", "multiplechoice"] and question.options:
            safe_options = [str(opt).replace('"', "'") for opt in question.options]
            formula = '"' + ','.join(safe_options) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            self.final_sheet.add_data_validation(dv)
            dv.add(answer_cell)
        elif question.type == "truefalse":
            dv = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
            self.final_sheet.add_data_validation(dv)
            dv.add(answer_cell)
        
        # Track for grading
        self.answer_cells.append({
            'question_id': str(question_num),
            'cell_address': answer_cell_addr,
            'correct_answer': question.answer,
            'tolerance': question.tolerance,
            'type': question.type
        })
        
        self.current_row += 1
        self.add_spacing(1)
        
        return answer_cell_addr

    def setup_grading(self):
        """Set up sophisticated grading formulas in Solution sheet"""
        grading_row = 2
        
        for answer_info in self.answer_cells:
            question_id = answer_info['question_id']
            cell_addr = answer_info['cell_address']
            correct_answer = answer_info['correct_answer']
            tolerance = answer_info.get('tolerance')
            question_type = answer_info.get('type', 'numerical')
            
            # Add to solution sheet
            self.solution_sheet.cell(row=grading_row, column=9, value=question_id)  # Column I
            
            # Student answer (reference to Student sheet) - CHANGED from "Final!" to "Student!"
            student_formula = f"=Student!{cell_addr}"
            self.solution_sheet.cell(row=grading_row, column=10, value=student_formula)  # Column J
            
            # Correct solution
            self.solution_sheet.cell(row=grading_row, column=11, value=correct_answer)  # Column K
            
            # Grading formula (Column L)
            if tolerance and question_type == "numerical":
                # Tolerance-based grading (like your manual exam!)
                grading_formula = f"=IF(ABS(VALUE(K{grading_row})-VALUE(J{grading_row}))<=({tolerance}*ABS(VALUE(K{grading_row}))),1,0)"
            elif question_type in ["truefalse", "multiplechoice", "dropdown"]:
                # Convert both to text for exact comparison (handles number vs string issue)
                grading_formula = f"=IF(TEXT(J{grading_row},\"@\")=TEXT(K{grading_row},\"@\"),1,0)"
            else:
                # Default numerical comparison with VALUE conversion
                grading_formula = f"=IF(ABS(VALUE(K{grading_row})-VALUE(J{grading_row}))<=0.01,1,0)"
            
            points_cell = self.solution_sheet.cell(row=grading_row, column=12, value=grading_formula)
            
            # Style the grading rows
            for col in range(9, 13):
                cell = self.solution_sheet.cell(row=grading_row, column=col)
                cell.border = self.thin_border
                cell.alignment = self.center_align
                if col == 9:  # Question ID column
                    cell.font = self.bold_font
            
            grading_row += 1
        
        # Add total score
        if self.answer_cells:
            total_cell = self.solution_sheet.cell(row=grading_row + 1, column=11, value="Total Score:")
            total_cell.font = self.bold_font
            total_cell.alignment = self.center_align
            
            total_formula = f"=SUM(L2:L{grading_row})"
            total_value_cell = self.solution_sheet.cell(row=grading_row + 1, column=12, value=total_formula)
            total_value_cell.font = self.bold_font
            total_value_cell.alignment = self.center_align
            total_value_cell.border = self.thin_border

def build_professional_workbook(questions: List[Question], filename="Professional_Exam.xlsx", exam_title="Professional Exam"):
    """Build a professional, beautiful Excel workbook like the manual example"""
    wb = Workbook()
    layout = ExcelLayoutEngine(wb)
    layout.setup_sheets()
    
    # Set custom title - ADDED: Use the exam_title parameter
    title_cell = layout.final_sheet.cell(row=1, column=2)
    title_cell.value = exam_title
    title_cell.font = layout.title_font
    title_cell.alignment = layout.center_align
    title_cell.fill = layout.white_fill
    
    question_num = 1
    
    for question in questions:
        if question.type == "data_table" or question.type == "multipart":
            # Complex question with data table and subquestions
            layout.add_question_header(question_num, question.question)
            
            if question.data_table:
                layout.add_data_table(question.data_table)
            
            if question.subquestions:
                for sub_q in question.subquestions:
                    layout.add_subquestion(sub_q, question_num)
            
        else:
            # Simple question
            layout.add_simple_question(question, question_num)
        
        question_num += 1
    
    # Set up grading system
    layout.setup_grading()
    
    # Hide solution sheet and protect final sheet
    layout.solution_sheet.sheet_state = 'hidden'
    layout.final_sheet.protection.sheet = False
    
    wb.save(filename)

@app.post("/ai-generate-question")
async def ai_generate_question(prompt: PromptRequest):
    try:
        if not prompt.prompt:
            raise HTTPException(status_code=400, detail="Prompt is required")

        # Enhanced system prompt for complex question generation
        system_prompt = """You are an AI that generates sophisticated exam questions with data tables. 

IMPORTANT: Return ONLY valid JSON. No markdown, no explanations, no extra text.

For BULK requests (multiple questions), return a JSON ARRAY like this:
[
 {
   "type": "data_table",
   "question": "Analyze the following GDP data:",
   "data_table": {
     "headers": ["Year", "GDP (Trillion $)"],
     "rows": [
       [2020, 21.4],
       [2021, 23.3],
       [2022, 25.5]
     ]
   },
   "subquestions": [
     {"part": "A", "question": "What was the GDP growth from 2020 to 2021?", "answer": "8.9", "tolerance": 0.02}
   ]
  }
]

- Make subquestions that require actual calculation
- Return ONLY the JSON, nothing else"""

        try:
            # Clean the input prompt first to remove problematic characters
            clean_prompt = prompt.prompt.encode('ascii', errors='ignore').decode('ascii')
            print(f"DEBUG: Original prompt length: {len(prompt.prompt)}")
            print(f"DEBUG: Cleaned prompt length: {len(clean_prompt)}")
            
            response = client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": clean_prompt}
                ],
                temperature=0.7,
                max_tokens=2000
            )
            print("DEBUG: OpenAI call successful")

            raw = response.choices[0].message.content.strip()
            print(f"DEBUG: Raw response length: {len(raw)}")
            
            # Handle encoding issues by cleaning the response
            raw = raw.encode('utf-8', errors='ignore').decode('utf-8')
            
            # More aggressive cleaning of the response
            # Remove markdown code blocks
            if "```json" in raw:
                parts = raw.split("```")
                if len(parts) >= 3:
                    raw = parts[1].strip()
                    if raw.startswith("json"):
                        raw = raw[4:].strip()

            # Remove any leading/trailing text that might not be JSON
            # Find the first { or [ and last } or ]
            start_idx = -1
            end_idx = -1
            
            for i, char in enumerate(raw):
                if char in ['{', '[']:
                    start_idx = i
                    break
            
            for i in range(len(raw) - 1, -1, -1):
                if raw[i] in ['}', ']']:
                    end_idx = i + 1
                    break
            
            if start_idx != -1 and end_idx != -1:
                raw = raw[start_idx:end_idx]
            
            print(f"DEBUG: Final cleaned JSON length: {len(raw)}")
            
            # Try to parse the JSON
            parsed = json.loads(raw)
            print("DEBUG: JSON parsing successful")
            return parsed

        except json.JSONDecodeError as e:
            print(f"DEBUG: JSON decode error: {str(e)}")
            print(f"DEBUG: Raw response: {raw[:500]}...")
            raise HTTPException(
                status_code=500,
                detail=f"AI returned invalid JSON. Error: {str(e)}"
            )
        except Exception as e:
            print(f"DEBUG: Other error in inner try: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Error in AI processing: {str(e)}"
            )
            
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except Exception as e:
        print(f"DEBUG: Error in outer try: {str(e)}")
        print(f"DEBUG: Error type: {type(e)}")
        import traceback
        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected error: {str(e)}"
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
